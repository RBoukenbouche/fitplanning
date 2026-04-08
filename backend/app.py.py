from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font
import datetime, json, io, os

app = Flask(__name__)
CORS(app)

BASE = os.path.dirname(__file__)

# ── MONTANT EN LETTRES ────────────────────────────────────────────────────────
ONES = ['','One','Two','Three','Four','Five','Six','Seven','Eight','Nine',
        'Ten','Eleven','Twelve','Thirteen','Fourteen','Fifteen','Sixteen',
        'Seventeen','Eighteen','Nineteen']
TENS = ['','','Twenty','Thirty','Forty','Fifty','Sixty','Seventy','Eighty','Ninety']

def _b1000(n):
    if n < 20: return ONES[n]
    elif n < 100: return TENS[n//10]+(' '+ONES[n%10] if ONES[n%10] else '')
    else: return ONES[n//100]+' Hundred'+(' '+_b1000(n%100) if n%100 else '')

def amount_to_words(amount, currency='EUR'):
    amount = round(amount, 2)
    main = int(amount); cents = round((amount - main) * 100)
    cur_name = 'Euros' if currency == 'EUR' else 'US Dollars'
    parts = []
    if main // 1_000_000_000: parts.append(_b1000(main//1_000_000_000)+' Billion')
    if (main%1_000_000_000)//1_000_000: parts.append(_b1000((main%1_000_000_000)//1_000_000)+' Million')
    if (main%1_000_000)//1_000: parts.append(_b1000((main%1_000_000)//1_000)+' Thousand')
    if main%1_000: parts.append(_b1000(main%1_000))
    result = ' '.join(parts) if parts else 'Zero'
    if cents: result += f' And {_b1000(cents)} Cents'
    return result + f' {cur_name} Only'

# ── SAFE SET ──────────────────────────────────────────────────────────────────
def s(ws, row, col, val):
    try: ws.cell(row=row, column=col).value = val
    except AttributeError: pass

def clr(ws, r1, r2, c1=1, c2=9):
    for r in range(r1, r2+1):
        for c in range(c1, c2+1): s(ws, r, c, None)

# ── FORMATAGE DATES ───────────────────────────────────────────────────────────
def fmt_short(date_str):
    """2026-04-23 → Apr 23"""
    try: return datetime.datetime.strptime(date_str, '%Y-%m-%d').strftime('%b %d')
    except: return date_str

def add_days(date_str, n):
    try:
        d = datetime.datetime.strptime(date_str, '%Y-%m-%d')
        return (d + datetime.timedelta(days=n)).strftime('%Y-%m-%d')
    except: return date_str

def fmt_period(arr, dep):
    """2026-04-23, 2026-05-01 → Apr 23 - May 01, 2026"""
    try:
        a = datetime.datetime.strptime(arr, '%Y-%m-%d')
        d = datetime.datetime.strptime(dep, '%Y-%m-%d')
        return f"{a.strftime('%b %d')} - {d.strftime('%b %d, %Y')}"
    except: return f"{arr} - {dep}"

# ── CONSTRUCTION LIGNES HÉBERGEMENT ──────────────────────────────────────────
def build_accom_lines(accom_data, arrival, num_days, num_cats):
    """
    Regroupe les nuits consécutives même hôtel/ville/type
    Format : Apr 23-25 : Marrakech - Riad Palais - Double Room - 2 units
    """
    lines = []
    for c in range(num_cats):
        entries = []
        for i in range(num_days):
            dd = accom_data.get(str(i), accom_data.get(i, {}))
            cd = dd.get(str(c), dd.get(c, {}))
            entries.append({
                'city':     cd.get('city', '').strip(),
                'hotel':    cd.get('hotel', '').strip(),
                'roomType': cd.get('roomType', '').strip(),
                'units':    int(float(cd.get('units', '0') or 0)),
                'day':      i
            })

        # Regrouper les nuits consécutives même hôtel
        groups = []
        cur = None
        for e in entries:
            key = (e['city'], e['hotel'], e['roomType'], e['units'])
            if not e['hotel'] and not e['city']:
                if cur: groups.append(cur); cur = None
                continue
            if cur and (cur['city'], cur['hotel'], cur['roomType'], cur['units']) == key:
                cur['end_day'] = e['day']
            else:
                if cur: groups.append(cur)
                cur = {**e, 'start_day': e['day'], 'end_day': e['day']}
        if cur: groups.append(cur)

        for g in groups:
            if not g['hotel'] and not g['city']: continue
            d_from = fmt_short(add_days(arrival, g['start_day']))
            d_to   = fmt_short(add_days(arrival, g['end_day'] + 1))
            parts  = [f"{d_from}-{d_to}"]
            if g['city']:     parts.append(g['city'])
            if g['hotel']:    parts.append(g['hotel'])
            if g['roomType']: parts.append(g['roomType'])
            u = g['units']
            if u > 0: parts.append(f"{u} unit{'s' if u > 1 else ''}")
            lines.append(' - '.join(parts))
    return lines

# ── CONSTRUCTION LIGNES MEALS ─────────────────────────────────────────────────
def build_meal_lines(meals_data):
    """
    Une ligne par repas, triée par date
    Format : Apr 23 : La Maison Arabe — Welcome Dinner
    """
    lines = []
    sorted_meals = sorted(meals_data, key=lambda x: x.get('date', ''))
    for m in sorted_meals:
        desc = m.get('desc', '').strip()
        date = m.get('date', '')
        if not desc: continue
        date_lbl = fmt_short(date) if date else ''
        lines.append(f"{date_lbl}: {desc}" if date_lbl else desc)
    return lines

# ── CONSTRUCTION LIGNES TRANSPORT ────────────────────────────────────────────
def build_trans_lines(trans_data, arrival, num_days):
    lines = []
    for i in range(num_days):
        row = trans_data.get(str(i), trans_data.get(i, []))
        desc = row[0].strip() if isinstance(row, list) and len(row) > 0 else ''
        if desc:
            lines.append(f"{fmt_short(add_days(arrival, i))}: {desc}")
    return lines

# ── CONSTRUCTION LIGNES ACTIVITÉS ────────────────────────────────────────────
def build_act_lines(act_data):
    lines = []
    if not act_data: return lines
    # Format dict {0: [desc, rate, qty]} venant de actBody
    if isinstance(act_data, dict):
        for i in sorted(act_data.keys(), key=lambda x: int(x) if str(x).isdigit() else 0):
            row = act_data[i]
            if isinstance(row, list) and len(row) > 0 and row[0]:
                lines.append(str(row[0]).strip())
            elif isinstance(row, dict):
                desc = row.get('desc', '').strip()
                if desc: lines.append(desc)
        return lines
    # Format liste [{date, desc}]
    if isinstance(act_data, list):
        for a in act_data:
            if not isinstance(a, dict): continue
            desc = a.get('desc', '').strip()
            date = a.get('date', '')
            if not desc: continue
            date_lbl = fmt_short(date) if date else ''
            lines.append(f"{date_lbl}: {desc}" if date_lbl else desc)
    return lines

# ── CONSTRUCTION LIGNES EXTRAS ────────────────────────────────────────────────
def build_extra_lines(extras_data):
    lines = []
    for e in extras_data:
        desc = e.get('desc', '').strip()
        if desc: lines.append(desc)
    return lines

# ── GÉNÉRATION FACTURE ────────────────────────────────────────────────────────
def generate_invoice(data):
    currency  = data.get('currency', 'USD')
    is_eur    = currency == 'EUR'
    pax       = int(data.get('pax', 1))
    total_mad = float(data.get('totalSell', 0))
    rate      = float(data.get('eurRate', 10.5))
    num_days  = int(data.get('numDays', 0))
    num_cats  = int(data.get('numCats', 1))
    arrival   = data.get('arr', '')
    departure = data.get('dep', '')
    total_f   = int(((total_mad / rate) + 9) // 10) * 10  # Arrondi dizaine superieure
    per_pax   = int(((total_f / pax) + 9) // 10) * 10 if pax > 0 else 0  # Arrondi dizaine superieure

    ref       = data.get('ref', '')
    client    = data.get('client', '')
    contact   = data.get('contact', '')
    pax_names = data.get('paxNames', '')
    agent     = data.get('agentDisplay', '')
    notes     = data.get('notes', '')
    today     = datetime.datetime.today()
    period    = fmt_period(arrival, departure)

    # Construire les lignes depuis les données brutes du devis
    accom_lines = build_accom_lines(data.get('accomData', {}), arrival, num_days, num_cats)
    meal_lines  = build_meal_lines(data.get('mealsData', []))
    trans_lines = build_trans_lines(data.get('transData', {}), arrival, num_days)
    act_lines   = build_act_lines(data.get('actData', []))
    extra_lines = build_extra_lines(data.get('extData', []))
    inc_lines   = data.get('incLines', [])
    if not inc_lines:
        inc_lines = [
            'Accommodation at hotels as shown above or Similar',
            'All meals as indicated',
            'A/C Private deluxe vehicle at disposal',
            'High qualified English Speaking guide throughout',
            'All visits and activities as per the Itinerary',
            'Entrance fees to the monuments',
            'Water in the car during the tour',
            'All Local Taxes',
        ]
    # total_f est le montant arrondi — words sera recalculé après insertion dans Excel
    words       = amount_to_words(total_f, currency)

    tpl_file = 'template_vim.xlsx' if is_eur else 'template_sweet.xlsx'
    tpl_path = os.path.join(BASE, tpl_file)
    wb = load_workbook(tpl_path)
    ws = wb.active

    if is_eur:
        # VIM — En-tête (lignes fixes 1-6 non touchées)
        s(ws, 9,  1, 'Bill To:');    s(ws, 10, 1, client);    s(ws, 11, 1, contact)
        s(ws, 9,  7, 'Date:');       s(ws, 9,  8, today)
        s(ws, 10, 7, 'Invoice N°:'); s(ws, 10, 8, ref)
        s(ws, 11, 7, 'Ref N°:');     s(ws, 11, 8, pax_names)
        s(ws, 12, 7, 'Ref:');        s(ws, 12, 8, period)

        # Package per pax (row 17) : Unit=pax, Days=num_days, Daily Rate=per_pax
        s(ws, 17, 1, f'Package per person — {period}')
        s(ws, 17, 5, pax)       # UNIT
        s(ws, 17, 6, num_days)  # DAYS
        s(ws, 17, 7, per_pax)   # DAILY RATE
        s(ws, 17, 8, '=G17*E17') # GROSS = Daily Rate × Unit

        clr(ws, 18, 48)
        row = 20

        if accom_lines:
            s(ws, row, 1, 'Accommodation:')
            ws.cell(row=row, column=1).font = Font(bold=True)
            row += 1
            for l in accom_lines: s(ws, row, 1, l); row += 1

        if meal_lines:
            s(ws, row, 1, 'Meals:')
            ws.cell(row=row, column=1).font = Font(bold=True)
            row += 1
            for l in meal_lines: s(ws, row, 1, l); row += 1

        if act_lines:
            s(ws, row, 1, 'Activities & Experiences:'); row += 1
            for l in act_lines: s(ws, row, 1, l); row += 1

        if extra_lines:
            s(ws, row, 1, 'Other Services & Extras:'); row += 1
            for l in extra_lines: s(ws, row, 1, l); row += 1

        if inc_lines:
            row += 1
            s(ws, row, 1, 'Including:')
            ws.cell(row=row, column=1).font = Font(bold=True)
            row += 1
            for l in inc_lines: s(ws, row, 1, f'- {l}'); row += 1

        if notes: row += 1; s(ws, row, 1, notes); row += 1

        tr = max(row + 2, 49)
        s(ws, tr,   5, f'Total Invoice in {currency}')
        s(ws, tr,   8, f'=SUM(H17:H{tr-1})')
        s(ws, tr+1, 7, f'Deposit for confirmation {currency}')
        s(ws, tr+1, 8, f'=0.3*H{tr}')
        s(ws, tr+2, 7, f'Balance to be paid {currency}')
        s(ws, tr+2, 8, f'=H{tr}-H{tr+1}')
        # Total en lettres = montant réel total facture
        total_final = total_f
        words_final = amount_to_words(total_final, currency)
        s(ws, tr+4, 1, words_final)

    else:
        # SWEET — En-tête
        s(ws, 7,  1, 'Bill To:');    s(ws, 8,  1, client);    s(ws, 9,  1, contact)
        s(ws, 7,  7, 'Date:');       s(ws, 7,  8, today)
        s(ws, 8,  7, 'Invoice N°:'); s(ws, 8,  8, ref)
        s(ws, 9,  7, 'Ref N°:');     s(ws, 9,  8, pax_names)
        s(ws, 10, 7, 'Ref:');        s(ws, 10, 8, period)

        # Package per pax (row 15)
        s(ws, 15, 1, f'Package per person — {period}')
        s(ws, 15, 2, pax)       # UNIT
        s(ws, 15, 3, num_days)  # DAYS
        s(ws, 15, 4, per_pax)   # DAILY RATE
        s(ws, 15, 5, 0)         # VAT 0
        s(ws, 15, 6, '=B15*C15*D15')
        s(ws, 15, 7, '=F15*E15')
        s(ws, 15, 8, '=F15+G15')

        clr(ws, 16, 44)
        row = 17

        if accom_lines:
            s(ws, row, 1, 'Accommodation:'); row += 1
            for l in accom_lines: s(ws, row, 1, l); row += 1

        if meal_lines:
            s(ws, row, 1, 'Meals:'); row += 1
            for l in meal_lines: s(ws, row, 1, l); row += 1

        if act_lines:
            s(ws, row, 1, 'Activities & Experiences:'); row += 1
            for l in act_lines: s(ws, row, 1, l); row += 1

        if extra_lines:
            s(ws, row, 1, 'Other Services & Extras:'); row += 1
            for l in extra_lines: s(ws, row, 1, l); row += 1

        if inc_lines:
            row += 1
            s(ws, row, 1, 'Including:')
            ws.cell(row=row, column=1).font = Font(bold=True)
            row += 1
            for l in inc_lines: s(ws, row, 1, f'- {l}'); row += 1

        if notes: row += 1; s(ws, row, 1, notes); row += 1

        tr = max(row + 2, 45)
        s(ws, tr,   4, 'Subtotal Excluding VAT')
        s(ws, tr,   8, f'=SUM(H15:H{tr-1})')
        s(ws, tr+1, 6, 'Total Output VAT')
        s(ws, tr+1, 8, '=G15')
        s(ws, tr+2, 5, f'Total Invoice in {currency}')
        s(ws, tr+2, 8, f'=SUM(H{tr}:H{tr+1})')
        s(ws, tr+3, 7, f'Deposit for confirmation {currency}')
        s(ws, tr+3, 8, f'=0.3*H{tr+2}')
        s(ws, tr+4, 7, f'Balance to be paid {currency}')
        s(ws, tr+4, 8, f'=H{tr+2}-H{tr+3}')
        # Total en lettres = montant réel total facture
        total_final = total_f
        words_final = amount_to_words(total_final, currency)
        s(ws, tr+6, 1, words_final)
        # Logo Sweet Spot
        logo_path = os.path.join(BASE, 'sweet_logo.png')
        if os.path.exists(logo_path):
            try:
                img = XLImage(logo_path)
                img.width = 150; img.height = 60
                ws.add_image(img, 'G1')
            except: pass

    # Sauvegarder en mémoire
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, words

# ── ROUTES ────────────────────────────────────────────────────────────────────
@app.route('/generate-invoice', methods=['POST'])
def route_invoice():
    try:
        data = request.get_json()
        buf, words = generate_invoice(data)
        currency = data.get('currency', 'USD')
        ref = data.get('ref', 'QT')
        client = data.get('client', 'Client').replace(' ', '_')
        filename = f"Invoice_{ref}_{client}.xlsx"
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok', 'message': 'Visit Morocco Invoice Server running'})

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5055, debug=False)

# ── EXPORT QUOTATION EXCEL ────────────────────────────────────────────────────
@app.route('/export-quotation', methods=['POST'])
def route_export_quotation():
    try:
        data = request.get_json()
        buf = generate_quotation_excel(data)
        ref = data.get('ref', 'QT')
        client = data.get('client', 'Client').replace(' ', '_')
        filename = f"Quotation_{ref}_{client}.xlsx"
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def generate_quotation_excel(data):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = data.get('ref', 'Quotation')

    # Couleurs
    navy = '1A2E4A'
    gold = 'C9A84C'
    light_blue = 'E8EFF8'
    green = '1A6B4A'
    white = 'FFFFFF'

    def hdr_style(cell, bg=navy, fg=white, bold=True, size=11):
        cell.font = Font(bold=bold, color=fg, size=size)
        cell.fill = PatternFill('solid', fgColor=bg)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    def sec_style(cell):
        cell.font = Font(bold=True, color=white, size=10)
        cell.fill = PatternFill('solid', fgColor=navy)
        cell.alignment = Alignment(horizontal='left', vertical='center')

    def gold_style(cell):
        cell.font = Font(bold=True, color=navy, size=10)
        cell.fill = PatternFill('solid', fgColor=gold)
        cell.alignment = Alignment(horizontal='right', vertical='center')

    thin = Border(
        left=Side(style='thin', color='D8E2EF'),
        right=Side(style='thin', color='D8E2EF'),
        top=Side(style='thin', color='D8E2EF'),
        bottom=Side(style='thin', color='D8E2EF')
    )

    def set_row(ws, row, values, styles=None):
        for i, val in enumerate(values):
            cell = ws.cell(row=row, column=i+1, value=val)
            cell.border = thin
            if styles and i < len(styles):
                st = styles[i]
                if st == 'hdr': hdr_style(cell)
                elif st == 'sec': sec_style(cell)
                elif st == 'gold': gold_style(cell)
                elif st == 'bold':
                    cell.font = Font(bold=True, size=10)
                elif st == 'num':
                    cell.alignment = Alignment(horizontal='right')
                    cell.number_format = '#,##0.00'

    # ── INFOS GÉNÉRALES ──────────────────────────────────────────────────────
    currency = data.get('currency', 'EUR')
    rate = float(data.get('rate', 10.5))
    pax = int(data.get('pax', 1))
    num_days = int(data.get('numDays', 0))
    num_cats = int(data.get('numCats', 1))
    arrival = data.get('arr', '')
    departure = data.get('dep', '')
    sell = lambda c, m: c/(1-(m/100)) if m < 100 else c

    r = 1
    ws.merge_cells(f'A{r}:H{r}')
    cell = ws.cell(row=r, column=1, value='VISIT MOROCCO TRAVEL & EVENTS — QUOTATION')
    hdr_style(cell, size=13)
    ws.row_dimensions[r].height = 28
    r += 1

    info_rows = [
        ('Réf. Devis', data.get('ref',''), 'Version', data.get('version',''), 'Statut', data.get('status',''), 'Date', data.get('quoteDate','')),
        ('Client', data.get('client',''), 'Contact', data.get('contact',''), 'Passagers', pax, 'Devise', currency),
        ('Arrivée', arrival, 'Départ', departure, 'Jours', num_days, 'Taux', f"1 {currency} = {rate} MAD"),
    ]
    for info in info_rows:
        for i, val in enumerate(info):
            cell = ws.cell(row=r, column=i+1, value=val)
            cell.border = thin
            if i % 2 == 0:
                cell.font = Font(bold=True, color='6B7C93', size=9)
                cell.fill = PatternFill('solid', fgColor=light_blue)
            else:
                cell.font = Font(size=10)
        r += 1
    r += 1

    # ── HÉBERGEMENT ──────────────────────────────────────────────────────────
    margin_accom = float(data.get('marginAccom', 0))
    ws.merge_cells(f'A{r}:H{r}')
    sec_style(ws.cell(row=r, column=1, value='🏨 HÉBERGEMENT'))
    ws.row_dimensions[r].height = 20
    r += 1

    set_row(ws, r, ['Jour','Date','Hôtel','Type Chambre','Unités','Tarif Unit. (MAD)','Total Achat (MAD)',''], 
            ['hdr','hdr','hdr','hdr','hdr','hdr','hdr','hdr'])
    r += 1

    accom_data = data.get('accomData', {})
    total_accom_cost = 0
    for i in range(num_days):
        for c in range(num_cats):
            row_d = (accom_data.get(str(i)) or accom_data.get(i) or {})
            cat_d = (row_d.get(str(c)) or row_d.get(c) or {})
            hotel = cat_d.get('hotel', '')
            room = cat_d.get('roomType', '')
            units = float(cat_d.get('units', 0) or 0)
            rate_u = float(cat_d.get('rate', 0) or 0)
            if not hotel: continue
            total = units * rate_u
            total_accom_cost += total
            date_lbl = fmt_short(add_days(arrival, i)) if arrival else f'Jour {i+1}'
            set_row(ws, r, [f'Jour {i+1}', date_lbl, hotel, room, units, rate_u, total, ''],
                    ['','','','','num','num','num',''])
            r += 1

    accom_sell = sell(total_accom_cost, margin_accom)
    set_row(ws, r, ['','','','','','Total Achat', total_accom_cost, ''],['','','','','','bold','num',''])
    r += 1
    set_row(ws, r, ['','','','','',f'Marge {margin_accom}%', '',''],['','','','','','bold','',''])
    r += 1
    ws.merge_cells(f'A{r}:F{r}')
    gold_style(ws.cell(row=r, column=1, value='TOTAL VENTE HÉBERGEMENT'))
    gold_style(ws.cell(row=r, column=7, value=accom_sell))
    ws.cell(row=r, column=7).number_format = '#,##0.00'
    r += 2

    # ── GUIDE & TRANSPORT ────────────────────────────────────────────────────
    margin_trans = float(data.get('marginTrans', 20))
    ws.merge_cells(f'A{r}:H{r}')
    sec_style(ws.cell(row=r, column=1, value='🚗 GUIDE & TRANSPORT'))
    ws.row_dimensions[r].height = 20
    r += 1

    set_row(ws, r, ['Jour','Date','Description','Véhicule (MAD)','Guide (MAD)','Qté','Total Achat (MAD)',''],
            ['hdr','hdr','hdr','hdr','hdr','hdr','hdr','hdr'])
    r += 1

    trans_data = data.get('transData', {})
    total_trans_cost = 0
    for i in range(num_days):
        row_t = trans_data.get(str(i)) or trans_data.get(i) or []
        desc = row_t[0] if len(row_t) > 0 else ''
        veh = float(row_t[1] or 0) if len(row_t) > 1 else 0
        guide = float(row_t[2] or 0) if len(row_t) > 2 else 0
        qty = float(row_t[3] or 1) if len(row_t) > 3 else 1
        if not desc: continue
        total = (veh + guide) * qty
        total_trans_cost += total
        date_lbl = fmt_short(add_days(arrival, i)) if arrival else f'Jour {i+1}'
        set_row(ws, r, [f'Jour {i+1}', date_lbl, desc, veh, guide, qty, total, ''],
                ['','','','num','num','num','num',''])
        r += 1

    trans_sell = sell(total_trans_cost, margin_trans)
    set_row(ws, r, ['','','','','','Total Achat', total_trans_cost, ''],['','','','','','bold','num',''])
    r += 1
    ws.merge_cells(f'A{r}:F{r}')
    gold_style(ws.cell(row=r, column=1, value='TOTAL VENTE TRANSPORT'))
    gold_style(ws.cell(row=r, column=7, value=trans_sell))
    ws.cell(row=r, column=7).number_format = '#,##0.00'
    r += 2

    # ── MEALS & SERVICES ─────────────────────────────────────────────────────
    margin_extras = float(data.get('marginExtras', 20))
    ws.merge_cells(f'A{r}:H{r}')
    sec_style(ws.cell(row=r, column=1, value='🍽️ MEALS & SERVICES'))
    ws.row_dimensions[r].height = 20
    r += 1

    set_row(ws, r, ['Description','Tarif Unit. (MAD)','Qté','Total Achat (MAD)','','','',''],
            ['hdr','hdr','hdr','hdr','','','',''])
    r += 1

    extras_data = data.get('extrasData', {})
    total_extras_cost = 0
    for val in extras_data.values():
        desc = val[0] if len(val) > 0 else ''
        tarif = float(val[1] or 0) if len(val) > 1 else 0
        qty = float(val[2] or 0) if len(val) > 2 else 0
        if not desc: continue
        total = tarif * qty
        total_extras_cost += total
        set_row(ws, r, [desc, tarif, qty, total, '','','',''],
                ['','num','num','num','','','',''])
        r += 1

    extras_sell = sell(total_extras_cost, margin_extras)
    set_row(ws, r, ['','Total Achat', total_extras_cost, '','','','',''],['','bold','num','','','','',''])
    r += 1
    ws.merge_cells(f'A{r}:C{r}')
    gold_style(ws.cell(row=r, column=1, value='TOTAL VENTE MEALS & SERVICES'))
    gold_style(ws.cell(row=r, column=4, value=extras_sell))
    ws.cell(row=r, column=4).number_format = '#,##0.00'
    r += 2

    # ── ACTIVITÉS ────────────────────────────────────────────────────────────
    margin_act = float(data.get('marginAct', 20))
    ws.merge_cells(f'A{r}:H{r}')
    sec_style(ws.cell(row=r, column=1, value='🎭 ACTIVITIES & EXPERIENCES'))
    ws.row_dimensions[r].height = 20
    r += 1

    set_row(ws, r, ['Jour','Date','Description','Tarif Unit. (MAD)','Qté','Total Achat (MAD)','',''],
            ['hdr','hdr','hdr','hdr','hdr','hdr','',''])
    r += 1

    act_data = data.get('actData', {})
    total_act_cost = 0
    for i in range(num_days):
        row_a = act_data.get(str(i)) or act_data.get(i) or []
        desc = row_a[0] if len(row_a) > 0 else ''
        tarif = float(row_a[1] or 0) if len(row_a) > 1 else 0
        qty = float(row_a[2] or 0) if len(row_a) > 2 else 0
        if not desc: continue
        total = tarif * qty
        total_act_cost += total
        date_lbl = fmt_short(add_days(arrival, i)) if arrival else f'Jour {i+1}'
        set_row(ws, r, [f'Jour {i+1}', date_lbl, desc, tarif, qty, total, '',''],
                ['','','','num','num','num','',''])
        r += 1

    act_sell = sell(total_act_cost, margin_act)
    set_row(ws, r, ['','','','','Total Achat', total_act_cost, '',''],['','','','','bold','num','',''])
    r += 1
    ws.merge_cells(f'A{r}:E{r}')
    gold_style(ws.cell(row=r, column=1, value='TOTAL VENTE ACTIVITÉS'))
    gold_style(ws.cell(row=r, column=6, value=act_sell))
    ws.cell(row=r, column=6).number_format = '#,##0.00'
    r += 2

    # ── RÉCAPITULATIF FINAL ──────────────────────────────────────────────────
    ws.merge_cells(f'A{r}:H{r}')
    sec_style(ws.cell(row=r, column=1, value='💰 RÉCAPITULATIF FINAL'))
    ws.row_dimensions[r].height = 20
    r += 1

    total_cost = total_accom_cost + total_trans_cost + total_extras_cost + total_act_cost
    total_sell_mad = accom_sell + trans_sell + extras_sell + act_sell
    total_sell_devise = int(((total_sell_mad / rate) + 9) // 10) * 10
    per_pax = int(((total_sell_mad / rate / pax) + 9) // 10) * 10

    recap = [
        ('🏨 Hébergement', f'{margin_accom}%', total_accom_cost, accom_sell, round(accom_sell/rate, 2)),
        ('🚗 Guide & Transport', f'{margin_trans}%', total_trans_cost, trans_sell, round(trans_sell/rate, 2)),
        ('🍽️ Meals & Services', f'{margin_extras}%', total_extras_cost, extras_sell, round(extras_sell/rate, 2)),
        ('🎭 Activités', f'{margin_act}%', total_act_cost, act_sell, round(act_sell/rate, 2)),
    ]
    set_row(ws, r, ['Section','Marge','Total Achat (MAD)','Total Vente (MAD)',f'Total Vente ({currency})','','',''],
            ['hdr','hdr','hdr','hdr','hdr','','',''])
    r += 1
    for rec in recap:
        set_row(ws, r, [rec[0], rec[1], rec[2], rec[3], rec[4],'','',''],
                ['','','num','num','num','','',''])
        r += 1

    # Total général
    set_row(ws, r, ['TOTAL GÉNÉRAL','', total_cost, total_sell_mad, total_sell_devise,'','',''],
            ['gold','gold','gold','gold','gold','','',''])
    for col in [3,4,5]:
        ws.cell(row=r, column=col).number_format = '#,##0.00'
    r += 1
    set_row(ws, r, [f'PRIX PAR PERSONNE ({pax} pax)','','','', per_pax,'','',''],
            ['gold','','','','gold','','',''])
    ws.cell(row=r, column=5).number_format = '#,##0.00'
    r += 1

    # Largeurs colonnes
    col_widths = [15, 12, 35, 18, 15, 12, 18, 10]
    for i, w in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(i+1)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
