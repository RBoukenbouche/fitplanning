from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
from openpyxl import load_workbook
from openpyxl.styles import Font
import datetime, json, io, os

app = Flask(__name__)
CORS(app)

BASE = os.path.dirname(__file__)

# ── MONTANT EN LETTRES ────────────────────────────────────────────────────────
ONES = ['','One','Two','Three','Four','Five','Six','Seven','Eight','Nine',
        'Ten','Eleven','Twelve','Thirteen','Fourteen','Fifteen','Sixteen',
        'Seventeen','Eighteen','Nineteen']
TENS = ['','','Twenty','Thirty','Forty','Fifty','Sixty','Seventy','Eighty','Ninety']

def _b1000(n):
    if n < 20: return ONES[n]
    elif n < 100: return TENS[n//10]+(' '+ONES[n%10] if ONES[n%10] else '')
    else: return ONES[n//100]+' Hundred'+(' '+_b1000(n%100) if n%100 else '')

def amount_to_words(amount, currency='EUR'):
    amount = round(amount, 2)
    main = int(amount); cents = round((amount - main) * 100)
    cur_name = 'Euros' if currency == 'EUR' else 'US Dollars'
    parts = []
    if main // 1_000_000_000: parts.append(_b1000(main//1_000_000_000)+' Billion')
    if (main%1_000_000_000)//1_000_000: parts.append(_b1000((main%1_000_000_000)//1_000_000)+' Million')
    if (main%1_000_000)//1_000: parts.append(_b1000((main%1_000_000)//1_000)+' Thousand')
    if main%1_000: parts.append(_b1000(main%1_000))
    result = ' '.join(parts) if parts else 'Zero'
    if cents: result += f' And {_b1000(cents)} Cents'
    return result + f' {cur_name} Only'

# ── SAFE SET ──────────────────────────────────────────────────────────────────
def s(ws, row, col, val):
    try: ws.cell(row=row, column=col).value = val
    except AttributeError: pass

def bold(ws, row, col):
    try: ws.cell(row=row, column=col).font = Font(bold=True)
    except AttributeError: pass

def clr(ws, r1, r2, c1=1, c2=9):
    for r in range(r1, r2+1):
        for c in range(c1, c2+1): s(ws, r, c, None)

# ── FORMATAGE DATES ───────────────────────────────────────────────────────────
def fmt_short(date_str):
    """2026-04-23 → Apr 23"""
    try: return datetime.datetime.strptime(date_str, '%Y-%m-%d').strftime('%b %d')
    except: return date_str

def add_days(date_str, n):
    try:
        d = datetime.datetime.strptime(date_str, '%Y-%m-%d')
        return (d + datetime.timedelta(days=n)).strftime('%Y-%m-%d')
    except: return date_str

def fmt_period(arr, dep):
    """2026-04-23, 2026-05-01 → Apr 23 - May 01, 2026"""
    try:
        a = datetime.datetime.strptime(arr, '%Y-%m-%d')
        d = datetime.datetime.strptime(dep, '%Y-%m-%d')
        return f"{a.strftime('%b %d')} - {d.strftime('%b %d, %Y')}"
    except: return f"{arr} - {dep}"

# ── CONSTRUCTION LIGNES HÉBERGEMENT ──────────────────────────────────────────
def build_accom_lines(accom_data, arrival, num_days, num_cats):
    """
    Regroupe les nuits consécutives même hôtel/ville/type
    Format : Apr 23-25 : Marrakech - Riad Palais - Double Room - 2 units
    """
    lines = []
    for c in range(num_cats):
        entries = []
        for i in range(num_days):
            dd = accom_data.get(str(i), accom_data.get(i, {}))
            cd = dd.get(str(c), dd.get(c, {}))
            entries.append({
                'city':     cd.get('city', '').strip(),
                'hotel':    cd.get('hotel', '').strip(),
                'roomType': cd.get('roomType', '').strip(),
                'units':    int(float(cd.get('units', '0') or 0)),
                'day':      i
            })

        # Regrouper les nuits consécutives même hôtel
        groups = []
        cur = None
        for e in entries:
            key = (e['city'], e['hotel'], e['roomType'], e['units'])
            if not e['hotel'] and not e['city']:
                if cur: groups.append(cur); cur = None
                continue
            if cur and (cur['city'], cur['hotel'], cur['roomType'], cur['units']) == key:
                cur['end_day'] = e['day']
            else:
                if cur: groups.append(cur)
                cur = {**e, 'start_day': e['day'], 'end_day': e['day']}
        if cur: groups.append(cur)

        for g in groups:
            if not g['hotel'] and not g['city']: continue
            d_from = fmt_short(add_days(arrival, g['start_day']))
            d_to   = fmt_short(add_days(arrival, g['end_day'] + 1))
            parts  = [f"{d_from}-{d_to}"]
            if g['city']:     parts.append(g['city'])
            if g['hotel']:    parts.append(g['hotel'])
            if g['roomType']: parts.append(g['roomType'])
            u = g['units']
            if u > 0: parts.append(f"{u} unit{'s' if u > 1 else ''}")
            lines.append(' - '.join(parts))
    return lines

# ── CONSTRUCTION LIGNES MEALS ─────────────────────────────────────────────────
def build_meal_lines(meals_data):
    """
    Une ligne par repas, triée par date
    Format : Apr 23 : La Maison Arabe — Welcome Dinner
    """
    lines = []
    sorted_meals = sorted(meals_data, key=lambda x: x.get('date', ''))
    for m in sorted_meals:
        desc = m.get('desc', '').strip()
        date = m.get('date', '')
        if not desc: continue
        date_lbl = fmt_short(date) if date else ''
        lines.append(f"{date_lbl}: {desc}" if date_lbl else desc)
    return lines

# ── CONSTRUCTION LIGNES TRANSPORT ────────────────────────────────────────────
def build_trans_lines(trans_data, arrival, num_days):
    lines = []
    for i in range(num_days):
        row = trans_data.get(str(i), trans_data.get(i, []))
        desc = row[0].strip() if isinstance(row, list) and len(row) > 0 else ''
        if desc:
            lines.append(f"{fmt_short(add_days(arrival, i))}: {desc}")
    return lines

# ── CONSTRUCTION LIGNES ACTIVITÉS ────────────────────────────────────────────
def build_act_lines(act_data):
    lines = []
    sorted_acts = sorted(act_data, key=lambda x: x.get('date', ''))
    for a in sorted_acts:
        desc = a.get('desc', '').strip()
        date = a.get('date', '')
        if not desc: continue
        date_lbl = fmt_short(date) if date else ''
        lines.append(f"{date_lbl}: {desc}" if date_lbl else desc)
    return lines

# ── CONSTRUCTION LIGNES EXTRAS ────────────────────────────────────────────────
def build_extra_lines(extras_data):
    lines = []
    for e in extras_data:
        desc = e.get('desc', '').strip()
        if desc: lines.append(desc)
    return lines

# ── GÉNÉRATION QUOTATION EXCEL (export-all-versions) ─────────────────────────
def generate_quotation_sheet(wb, v):
    from openpyxl.styles import PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    NAVY   = 'FF1A2E4A'
    GOLD   = 'FFC9A84C'
    BLUE_L = 'FFE8EFF8'
    WHITE  = 'FFFFFFFF'

    def hdr_fill(): return PatternFill('solid', fgColor=NAVY)
    def gold_fill(): return PatternFill('solid', fgColor=GOLD)
    def blue_fill(): return PatternFill('solid', fgColor=BLUE_L)

    def hdr_font(size=10): return Font(bold=True, color='FFFFFFFF', size=size)
    def bold_font(size=10): return Font(bold=True, size=size)
    def gold_bold(): return Font(bold=True, size=10)

    def center(): return Alignment(horizontal='center')
    def right():  return Alignment(horizontal='right')
    def left():   return Alignment(horizontal='left')

    sheet_name = f"{v.get('ref','QT')}-{v.get('version','v1')}"[:31]
    ws = wb.create_sheet(title=sheet_name)

    rate        = float(v.get('rate', 10.5))
    pax         = int(v.get('pax', 1))
    num_days    = int(v.get('numDays', 0))
    num_cats    = int(v.get('numCats', 1))
    arrival     = v.get('arr', '')
    departure   = v.get('dep', '')
    currency    = v.get('currency', 'EUR')
    margin_a    = float(v.get('marginAccom', 0))
    margin_t    = float(v.get('marginTrans', 20))
    margin_x    = float(v.get('marginAct', 20))
    margin_e    = float(v.get('marginExtras', 20))
    accom_data  = v.get('accomData', {})
    trans_data  = v.get('transData', {})
    act_data    = v.get('actData', [])
    extras_data = v.get('extrasData', {})

    # Column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 20

    # ── ROW 1 : Title ────────────────────────────────────────────────────────
    ws.merge_cells('A1:H1')
    ws['A1'] = 'VISIT MOROCCO TRAVEL & EVENTS — QUOTATION'
    ws['A1'].font = Font(bold=True, color='FFFFFFFF', size=14)
    ws['A1'].fill = hdr_fill()
    ws['A1'].alignment = Alignment(horizontal='center')

    # ── ROW 2 : Ref / Version / Statut / Date ────────────────────────────────
    for col, label, val in [(1,'Réf. Devis',v.get('ref','')), (3,'Version',v.get('version','v1')),
                             (5,'Statut',v.get('status','draft')), (7,'Date',v.get('quoteDate',''))]:
        ws.cell(2, col, label).font = bold_font()
        ws.cell(2, col).fill = blue_fill()
        ws.cell(2, col+1, val)

    # ── ROW 3 : Client / Contact / Pax / Devise ──────────────────────────────
    ws.cell(3,1,'Client').font = bold_font(); ws.cell(3,1).fill = blue_fill()
    ws.cell(3,2, v.get('client',''))
    ws.cell(3,3,'Contact').font = bold_font(); ws.cell(3,3).fill = blue_fill()
    ws.cell(3,4, v.get('contact',''))
    ws.cell(3,5,'Passagers').font = bold_font(); ws.cell(3,5).fill = blue_fill()
    ws.cell(3,6, pax)
    ws.cell(3,7,'Devise').font = bold_font(); ws.cell(3,7).fill = blue_fill()
    ws.cell(3,8, currency)

    # ── ROW 4 : Arrivée / Départ / Jours / Taux ──────────────────────────────
    ws.cell(4,1,'Arrivée').font = bold_font(); ws.cell(4,1).fill = blue_fill()
    ws.cell(4,2, arrival)
    ws.cell(4,3,'Départ').font = bold_font(); ws.cell(4,3).fill = blue_fill()
    ws.cell(4,4, departure)
    ws.cell(4,5,'Jours').font = bold_font(); ws.cell(4,5).fill = blue_fill()
    ws.cell(4,6, num_days)
    ws.cell(4,7,'Taux').font = bold_font(); ws.cell(4,7).fill = blue_fill()
    ws.cell(4,8, rate)

    cur_row = 6

    # ── HÉBERGEMENT ──────────────────────────────────────────────────────────
    ws.merge_cells(f'A{cur_row}:H{cur_row}')
    ws.cell(cur_row,1, f'🏨 HEBERGEMENT  (Marge: {margin_a}%)').font = hdr_font(11)
    ws.cell(cur_row,1).fill = hdr_fill()
    ws.cell(cur_row,1).alignment = left()
    cur_row += 1

    for col, lbl in enumerate(['Jour','Date','Hotel','Type Chambre','Unites','Tarif Unit. (MAD)','Total Achat (MAD)','Total Vente (MAD)'],1):
        ws.cell(cur_row, col, lbl).font = hdr_font()
        ws.cell(cur_row, col).fill = hdr_fill()
        ws.cell(cur_row, col).alignment = center()
    cur_row += 1

    accom_start = cur_row
    for i in range(num_days):
        for c in range(num_cats):
            dd = accom_data.get(str(i), accom_data.get(i, {}))
            cd = dd.get(str(c), dd.get(c, {}))
            hotel = cd.get('hotel','').strip()
            city  = cd.get('city','').strip()
            room  = cd.get('roomType','').strip()
            units = int(float(cd.get('units',0) or 0))
            rate_v = float(cd.get('rate',0) or 0)
            if not hotel and not city: continue
            date_lbl = fmt_short(add_days(arrival, i)) if arrival else f'J{i+1}'
            ws.cell(cur_row,1, f'Jour {i+1}')
            ws.cell(cur_row,2, date_lbl)
            ws.cell(cur_row,3, f'{city} - {hotel}' if city else hotel)
            ws.cell(cur_row,4, room)
            ws.cell(cur_row,5, units)
            ws.cell(cur_row,6, rate_v)
            ws.cell(cur_row,7, f'=E{cur_row}*F{cur_row}')
            ws.cell(cur_row,8, f'=CEILING(G{cur_row}/(1-{margin_a/100}),10)' if margin_a > 0 else f'=G{cur_row}')
            cur_row += 1

    accom_end = cur_row - 1
    total_row_a = cur_row
    ws.cell(total_row_a, 5, 'TOTAL').font = bold_font()
    ws.cell(total_row_a, 7, f'=SUM(G{accom_start}:G{accom_end})').font = bold_font()
    ws.cell(total_row_a, 7).fill = gold_fill(); ws.cell(total_row_a, 7).alignment = right()
    ws.cell(total_row_a, 8, f'=SUM(H{accom_start}:H{accom_end})').font = bold_font()
    ws.cell(total_row_a, 8).fill = gold_fill(); ws.cell(total_row_a, 8).alignment = right()
    cur_row += 2

    # ── GUIDE & TRANSPORT ────────────────────────────────────────────────────
    ws.merge_cells(f'A{cur_row}:H{cur_row}')
    ws.cell(cur_row,1, f'🚗 GUIDE & TRANSPORT  (Marge: {margin_t}%)').font = hdr_font(11)
    ws.cell(cur_row,1).fill = hdr_fill(); ws.cell(cur_row,1).alignment = left()
    cur_row += 1

    for col, lbl in enumerate(['Jour','Date','Description','Vehicule (MAD)','Guide (MAD)','Qte','Total Achat (MAD)','Total Vente (MAD)'],1):
        ws.cell(cur_row, col, lbl).font = hdr_font()
        ws.cell(cur_row, col).fill = hdr_fill()
        ws.cell(cur_row, col).alignment = center()
    cur_row += 1

    trans_start = cur_row
    for i in range(num_days + 1):
        row_t = trans_data.get(str(i), trans_data.get(i, []))
        desc  = row_t[0].strip() if isinstance(row_t, list) and len(row_t) > 0 and row_t[0] else ''
        veh   = float(row_t[1] or 0) if isinstance(row_t, list) and len(row_t) > 1 else 0
        guide = float(row_t[2] or 0) if isinstance(row_t, list) and len(row_t) > 2 else 0
        qty   = float(row_t[3] or 1) if isinstance(row_t, list) and len(row_t) > 3 else 1
        date_lbl = fmt_short(add_days(arrival, i)) if arrival else f'J{i+1}'
        ws.cell(cur_row,1, f'Jour {i+1}')
        ws.cell(cur_row,2, date_lbl)
        ws.cell(cur_row,3, desc)
        ws.cell(cur_row,4, veh)
        ws.cell(cur_row,5, guide)
        ws.cell(cur_row,6, qty)
        ws.cell(cur_row,7, f'=(D{cur_row}+E{cur_row})*F{cur_row}')
        ws.cell(cur_row,8, f'=CEILING(G{cur_row}/(1-{margin_t/100}),10)')
        cur_row += 1

    trans_end = cur_row - 1
    total_row_t = cur_row
    ws.cell(total_row_t, 5, 'TOTAL').font = bold_font()
    ws.cell(total_row_t, 7, f'=SUM(G{trans_start}:G{trans_end})').font = bold_font()
    ws.cell(total_row_t, 7).fill = gold_fill(); ws.cell(total_row_t, 7).alignment = right()
    ws.cell(total_row_t, 8, f'=SUM(H{trans_start}:H{trans_end})').font = bold_font()
    ws.cell(total_row_t, 8).fill = gold_fill(); ws.cell(total_row_t, 8).alignment = right()
    cur_row += 2

    # ── MEALS & SERVICES ────────────────────────────────────────────────────
    ws.merge_cells(f'A{cur_row}:H{cur_row}')
    ws.cell(cur_row,1, f'🍽️ MEALS & SERVICES  (Marge: {margin_x}%)').font = hdr_font(11)
    ws.cell(cur_row,1).fill = hdr_fill(); ws.cell(cur_row,1).alignment = left()
    cur_row += 1

    for col, lbl in enumerate(['Description','Tarif Unit. (MAD)','Qte','Total Achat (MAD)','Total Vente (MAD)'],1):
        ws.cell(cur_row, col, lbl).font = hdr_font()
        ws.cell(cur_row, col).fill = hdr_fill()
        ws.cell(cur_row, col).alignment = center()
    cur_row += 1

    meals_start = cur_row
    for a in act_data:
        desc = a.get('desc','').strip()
        rate_v = float(a.get('rate',0) or 0)
        qty  = float(a.get('qty',0) or 0)
        if not desc: continue
        ws.cell(cur_row,1, desc)
        ws.cell(cur_row,2, rate_v)
        ws.cell(cur_row,3, qty)
        ws.cell(cur_row,4, f'=B{cur_row}*C{cur_row}')
        ws.cell(cur_row,5, f'=CEILING(D{cur_row}/(1-{margin_x/100}),10)')
        cur_row += 1

    meals_end = cur_row - 1
    total_row_m = cur_row
    ws.cell(total_row_m, 2, 'TOTAL').font = bold_font()
    ws.cell(total_row_m, 4, f'=SUM(D{meals_start}:D{meals_end})').font = bold_font()
    ws.cell(total_row_m, 4).fill = gold_fill(); ws.cell(total_row_m, 4).alignment = right()
    ws.cell(total_row_m, 5, f'=SUM(E{meals_start}:E{meals_end})').font = bold_font()
    ws.cell(total_row_m, 5).fill = gold_fill(); ws.cell(total_row_m, 5).alignment = right()
    cur_row += 2

    # ── RECAP FINAL ─────────────────────────────────────────────────────────
    ws.merge_cells(f'A{cur_row}:H{cur_row}')
    ws.cell(cur_row,1, '💰 RECAP FINAL').font = hdr_font(11)
    ws.cell(cur_row,1).fill = hdr_fill(); ws.cell(cur_row,1).alignment = left()
    cur_row += 1

    for col, lbl in enumerate(['Section','Marge %',f'Total Achat (MAD)',f'Total Vente (MAD)',f'Total Vente ({currency})'],1):
        ws.cell(cur_row, col, lbl).font = hdr_font()
        ws.cell(cur_row, col).fill = hdr_fill()
        ws.cell(cur_row, col).alignment = center()
    cur_row += 1

    recap_start = cur_row
    ws.cell(cur_row,1,'Hebergement'); ws.cell(cur_row,2,f'{margin_a}%')
    ws.cell(cur_row,3,f'=G{total_row_a}'); ws.cell(cur_row,4,f'=H{total_row_a}')
    ws.cell(cur_row,5,f'=CEILING(D{cur_row}/H4,10)')
    cur_row += 1
    ws.cell(cur_row,1,'Transport'); ws.cell(cur_row,2,f'{margin_t}%')
    ws.cell(cur_row,3,f'=G{total_row_t}'); ws.cell(cur_row,4,f'=H{total_row_t}')
    ws.cell(cur_row,5,f'=CEILING(D{cur_row}/H4,10)')
    cur_row += 1
    ws.cell(cur_row,1,'Meals & Services'); ws.cell(cur_row,2,f'{margin_x}%')
    ws.cell(cur_row,3,f'=D{total_row_m}'); ws.cell(cur_row,4,f'=E{total_row_m}')
    ws.cell(cur_row,5,f'=CEILING(D{cur_row}/H4,10)')
    recap_end = cur_row
    cur_row += 1

    total_row_f = cur_row
    ws.cell(cur_row,1,'TOTAL GENERAL').font = bold_font()
    ws.cell(cur_row,1).fill = gold_fill(); ws.cell(cur_row,1).alignment = right()
    ws.cell(cur_row,3,f'=SUM(C{recap_start}:C{recap_end})').font = bold_font()
    ws.cell(cur_row,3).fill = gold_fill(); ws.cell(cur_row,3).alignment = right()
    ws.cell(cur_row,4,f'=SUM(D{recap_start}:D{recap_end})').font = bold_font()
    ws.cell(cur_row,4).fill = gold_fill(); ws.cell(cur_row,4).alignment = right()
    ws.cell(cur_row,5,f'=CEILING(SUM(D{recap_start}:D{recap_end})/H4,10)').font = bold_font()
    ws.cell(cur_row,5).fill = gold_fill(); ws.cell(cur_row,5).alignment = right()
    cur_row += 1

    ws.cell(cur_row,1,f'PRIX PAR PERSONNE ({pax} pax)').font = bold_font()
    ws.cell(cur_row,1).fill = gold_fill(); ws.cell(cur_row,1).alignment = right()
    ws.cell(cur_row,5,f'=CEILING(E{total_row_f}/{pax},10)').font = bold_font()
    ws.cell(cur_row,5).fill = gold_fill(); ws.cell(cur_row,5).alignment = right()

# ── GÉNÉRATION FACTURE ────────────────────────────────────────────────────────
def generate_invoice(data):
    currency  = data.get('currency', 'USD')
    is_eur    = currency == 'EUR'
    pax       = int(data.get('pax', 1))
    total_mad = float(data.get('totalSell', 0))
    rate      = float(data.get('eurRate', 10.5))
    num_days  = int(data.get('numDays', 0))
    num_cats  = int(data.get('numCats', 1))
    arrival   = data.get('arr', '')
    departure = data.get('dep', '')
    total_f   = round(total_mad / rate, 2)
    per_pax   = round(total_f / pax, 2) if pax > 0 else 0

    ref       = data.get('ref', '')
    client    = data.get('client', '')
    contact   = data.get('contact', '')
    pax_names = data.get('paxNames', '')
    agent     = data.get('agentDisplay', '')
    notes     = data.get('notes', '')
    today     = datetime.datetime.today()
    period    = fmt_period(arrival, departure)

    # Construire les lignes depuis les données brutes du devis
    accom_lines = build_accom_lines(data.get('accomData', {}), arrival, num_days, num_cats)
    meal_lines  = build_meal_lines(data.get('mealsData', []))
    trans_lines = build_trans_lines(data.get('transData', {}), arrival, num_days)
    act_lines   = build_act_lines(data.get('actData', []))
    extra_lines = build_extra_lines(data.get('extData', []))
    inc_lines   = data.get('incLines', [])
    words       = amount_to_words(total_f, currency)

    tpl_file = 'template_vim.xlsx' if is_eur else 'template_sweet.xlsx'
    tpl_path = os.path.join(BASE, tpl_file)
    wb = load_workbook(tpl_path)
    ws = wb.active

    if is_eur:
        # VIM — En-tête (lignes fixes 1-6 non touchées)
        s(ws, 9,  1, 'Bill To:');    s(ws, 10, 1, client);    s(ws, 11, 1, contact)
        s(ws, 9,  7, 'Date:');       s(ws, 9,  8, today)
        s(ws, 10, 7, 'Invoice N°:'); s(ws, 10, 8, ref)
        s(ws, 11, 7, 'Ref N°:');     s(ws, 11, 8, pax_names)
        s(ws, 12, 7, 'Ref:');        s(ws, 12, 8, period)

        # Package per pax (row 17) : Unit=pax, Days=num_days, Daily Rate=per_pax
        s(ws, 17, 1, f'Package per person — {period}')
        s(ws, 17, 5, pax)       # UNIT
        s(ws, 17, 6, num_days)  # DAYS
        s(ws, 17, 7, per_pax)   # DAILY RATE
        s(ws, 17, 8, '=G17*E17') # GROSS = Daily Rate × Unit

        clr(ws, 18, 48)
        row = 20

        if accom_lines:
            s(ws, row, 1, 'Accommodation:'); bold(ws, row, 1); row += 1
            for l in accom_lines: s(ws, row, 1, l); row += 1

        if meal_lines:
            s(ws, row, 1, 'Meals & Experiences:'); bold(ws, row, 1); row += 1
            for l in meal_lines: s(ws, row, 1, l); row += 1

        if trans_lines:
            s(ws, row, 1, 'Guide & Transportation:'); row += 1
            for l in trans_lines: s(ws, row, 1, l); row += 1

        if act_lines:
            s(ws, row, 1, 'Activities & Experiences:'); row += 1
            for l in act_lines: s(ws, row, 1, l); row += 1

        if inc_lines:
            row += 1; s(ws, row, 1, 'Including:'); row += 1
            for l in inc_lines: s(ws, row, 1, f'- {l}'); row += 1

        if notes: row += 1; s(ws, row, 1, notes); row += 1

        tr = max(row + 2, 49)
        s(ws, tr,   5, f'Total Invoice in {currency}')
        s(ws, tr,   8, f'=SUM(H17:H{tr-1})')
        s(ws, tr+1, 7, f'Deposit for confirmation {currency}')
        s(ws, tr+1, 8, f'=0.3*H{tr}')
        s(ws, tr+2, 7, f'Balance to be paid {currency}')
        s(ws, tr+2, 8, f'=H{tr}-H{tr+1}')
        s(ws, tr+4, 1, words)

    else:
        # SWEET — En-tête
        s(ws, 7,  1, 'Bill To:');    s(ws, 8,  1, client);    s(ws, 9,  1, contact)
        s(ws, 7,  7, 'Date:');       s(ws, 7,  8, today)
        s(ws, 8,  7, 'Invoice N°:'); s(ws, 8,  8, ref)
        s(ws, 9,  7, 'Ref N°:');     s(ws, 9,  8, pax_names)
        s(ws, 10, 7, 'Ref:');        s(ws, 10, 8, period)

        # Package per pax (row 15)
        s(ws, 15, 1, f'Package per person — {period}')
        s(ws, 15, 2, pax)       # UNIT
        s(ws, 15, 3, num_days)  # DAYS
        s(ws, 15, 4, per_pax)   # DAILY RATE
        s(ws, 15, 5, 0)         # VAT 0
        s(ws, 15, 6, '=B15*D15')
        s(ws, 15, 7, '=F15*E15')
        s(ws, 15, 8, '=F15+G15')

        clr(ws, 16, 44)
        row = 17

        if accom_lines:
            s(ws, row, 1, 'Accommodation:'); bold(ws, row, 1); row += 1
            for l in accom_lines: s(ws, row, 1, l); row += 1

        if meal_lines:
            s(ws, row, 1, 'Meals & Experiences:'); bold(ws, row, 1); row += 1
            for l in meal_lines: s(ws, row, 1, l); row += 1

        if trans_lines:
            s(ws, row, 1, 'Guide & Transportation:'); row += 1
            for l in trans_lines: s(ws, row, 1, l); row += 1

        if act_lines:
            s(ws, row, 1, 'Activities & Experiences:'); row += 1
            for l in act_lines: s(ws, row, 1, l); row += 1

        if inc_lines:
            row += 1; s(ws, row, 1, 'Including:'); row += 1
            for l in inc_lines: s(ws, row, 1, f'- {l}'); row += 1

        if notes: row += 1; s(ws, row, 1, notes); row += 1

        tr = max(row + 2, 45)
        s(ws, tr,   4, 'Subtotal Excluding VAT')
        s(ws, tr,   8, f'=SUM(H15:H{tr-1})')
        s(ws, tr+1, 6, 'Total Output VAT')
        s(ws, tr+1, 8, '=G15')
        s(ws, tr+2, 5, f'Total Invoice in {currency}')
        s(ws, tr+2, 8, f'=SUM(H{tr}:H{tr+1})')
        s(ws, tr+3, 7, f'Deposit for confirmation {currency}')
        s(ws, tr+3, 8, f'=0.3*H{tr+2}')
        s(ws, tr+4, 7, f'Balance to be paid {currency}')
        s(ws, tr+4, 8, f'=H{tr+2}-H{tr+3}')
        s(ws, tr+6, 1, words)

    # Sauvegarder en mémoire
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, words

# ── ROUTES ────────────────────────────────────────────────────────────────────
@app.route('/generate-invoice', methods=['POST'])
def route_invoice():
    try:
        data = request.get_json()
        buf, words = generate_invoice(data)
        currency = data.get('currency', 'USD')
        ref = data.get('ref', 'QT')
        client = data.get('client', 'Client').replace(' ', '_')
        filename = f"Invoice_{ref}_{client}.xlsx"
        return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=filename)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/export-all-versions', methods=['POST', 'HEAD', 'GET'])
def route_export_all():
    if request.method in ('HEAD', 'GET'):
        return '', 200
    try:
        from openpyxl import Workbook
        data = request.get_json()
        versions = data.get('versions', [])
        if not versions:
            return jsonify({'error': 'No versions provided'}), 400
        wb = Workbook()
        wb.remove(wb.active)
        for v in versions:
            generate_quotation_sheet(wb, v)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        ref = versions[0].get('ref', 'QT') if versions else 'QT'
        client = versions[0].get('client', 'Client').replace(' ', '_') if versions else 'Client'
        filename = f"Quotation_{ref}_{client}_AllVersions.xlsx"
        return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=filename)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok', 'message': 'Visit Morocco Invoice Server running'})

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5055, debug=False)
